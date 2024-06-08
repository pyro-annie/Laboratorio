from flask import Flask, request, jsonify, render_template
import re
import secrets
import string
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)

def load_or_create_workbook(file_path):
    if os.path.exists(file_path):
        return load_workbook(file_path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre de Usuario', 'Contraseña', 'Correo Electrónico', 'Número de Teléfono'])
        wb.save(file_path)
        return wb

headers = ['Nombre de Usuario', 'Contraseña', 'Correo Electrónico', 'Número de Teléfono']
file_path = 'cuentas.xlsx'
wb = load_or_create_workbook(file_path)
ws = wb.active

def get_user_by_phone(phone_number, file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if phone_number == row[3]:
            return {'username': row[0], 'password': row[1], 'email': row[2], 'phone': row[3]}
    return None

EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@(gmail\.com|outlook\.com)$')

password_requirements = {
    'uppercase': 1,
    'digits': 1,
    'special': 1,
    'min_length': 8
}

def check_password_requirements(password):
    patterns = {
        'uppercase': r'[A-Z]',
        'digits': r'\d',
        'special': r'\W'
    }
    return all(len(re.findall(pattern, password)) >= password_requirements[requirement]
               for requirement, pattern in patterns.items()) and len(password) >= password_requirements['min_length']

def generate_secure_password():
    alphabet = string.ascii_letters + string.digits + string.punctuation
    while True:
        password = ''.join(secrets.choice(alphabet) for i in range(password_requirements['min_length']))
        if check_password_requirements(password):
            break
    return password

def register_user(username, password, email, phone, file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
        if username in row:
            return 'El nombre de usuario ya está en uso. Por favor, elige otro.'
    for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, values_only=True):
        if email in row:
            return 'Correo Electrónico ya en uso. Por favor, utiliza otro correo.'
    ws.append([username, password, email, phone])
    wb.save(file_path)
    return 'Usuario registrado con éxito.'

def is_phone_registered(phone_number, file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if phone_number == row[3]:
            return True
    return False

def get_user_by_username(username, file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if username == row[0]:
            return {'username': row[0], 'password': row[1]}
    return None

@app.route('/verify-phone', methods=['POST'])
def verify_phone():
    phone_number = request.form['phone'].strip()
    user = get_user_by_phone(phone_number, 'cuentas.xlsx')
    if user:
        return render_template('enter_code.html')
    else:
        return jsonify({'error': 'Número de teléfono no registrado.'}), 400

@app.route('/generate-password', methods=['POST'])
def generate_password():
    return jsonify({'password': generate_secure_password()})

@app.route('/forgot-account')
def forgot_account():
    return render_template('forgotaccount.html')

@app.route('/register', methods=['POST'])
def register():
    name = request.form['name'].strip()
    password = request.form['password']
    phone = request.form['phone'].strip()
    email = request.form['email'].strip()
    file_path = 'cuentas.xlsx'

    # Cargar el libro de trabajo de Excel
    wb = load_workbook(file_path)
    ws = wb.active

    # Verificar si el nombre de usuario ya está registrado
    for row in ws.iter_rows(min_row=2, values_only=True):
        if name == row[0]:
            return jsonify({'error': 'El nombre de usuario ya está en uso. Por favor, elige otro.'}), 400

    # Verificar los requisitos de la contraseña
    if not check_password_requirements(password):
        return jsonify({'error': 'La contraseña no cumple con los requisitos necesarios.'}), 400

    # Verificar que todos los campos estén completos
    if not name or not password or not phone or not email:
        return jsonify({'error': 'Todos los campos son obligatorios.'}), 400

    # Verificar el formato del correo electrónico
    if not EMAIL_REGEX.match(email):
        return jsonify({'error': 'Por favor, ingresa un correo electrónico válido de Gmail o Outlook.'}), 400

    # Verificar si el correo electrónico ya está en uso
    for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, values_only=True):
        if email == row[2]:
            return jsonify({'error': 'Correo Electrónico ya en uso. Por favor, utiliza otro correo.'}), 400

    # Agregar el nuevo usuario al libro de Excel
    ws.append([name, password, email, phone])
    wb.save(file_path)

    return jsonify({'message': 'Usuario registrado con éxito.'})

@app.route('/login', methods=['POST'])
def login():
    username = request.form['username'].strip()
    password = request.form['password']
    file_path = 'cuentas.xlsx'
    user = get_user_by_username(username, file_path)
    if not user:
        return jsonify({'error': 'El nombre de usuario no está registrado. Por favor, verifica tus datos o regístrate.'}), 400
    if user['password'] != password:
        return jsonify({'error': 'La contraseña introducida es incorrecta. Por favor, inténtalo de nuevo.'}), 400
    return jsonify({'message': f'Bienvenido, {username}'})

@app.route('/')
def index():
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
